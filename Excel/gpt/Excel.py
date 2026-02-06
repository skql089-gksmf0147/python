import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string
from datetime import datetime
import pandas as pd
import os
import shutil
import tkinter.font as tkFont
import warnings

warnings.filterwarnings("ignore", category=UserWarning)

# ================= 공통 설정 및 함수 =================
DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")

def find_excel(machine_name):
    for f in os.listdir(DESKTOP):
        if f.endswith(".xlsx") and machine_name in f:
            return os.path.join(DESKTOP, f)
    return None

def prev_month():
    now = datetime.now()
    y, m = now.year, now.month
    if m == 0: y -= 1; m = 12
    return y, m

def run_backup(machine_name):
    path = find_excel(machine_name)
    if not path:
        messagebox.showerror("에러", f"{machine_name} 엑셀을 찾지 못했습니다.")
        return

    wb = load_workbook(path)
    day_sheets = [int(s) for s in wb.sheetnames if s.isdigit()]
    if not day_sheets:
        messagebox.showerror("에러", "날짜 시트가 없습니다.")
        return

    year, month = prev_month()
    backup_name = f"백업_{year}_{month:02d}"

    if backup_name in wb.sheetnames:
        messagebox.showwarning("중복", f"{backup_name} 시트가 이미 존재합니다.")
        return

    backup_ws = wb.create_sheet(backup_name)
    wb._sheets.remove(backup_ws)
    wb._sheets.append(backup_ws)
    backup_ws["A1"] = f"{year}년 {month}월 백업"
    backup_ws["A1"].font = Font(bold=True)

    backup_row = 3
    for day in sorted(day_sheets):
        ws = wb[str(day)]
        rows = [r for r in range(9, ws.max_row + 1) if ws[f"A{r}"].value in ("정미시간", "준비시간")]
        if not rows: continue

        backup_ws.merge_cells(start_row=backup_row, start_column=1, end_row=backup_row, end_column=21)
        backup_ws.cell(row=backup_row, column=1).value = f"{month}월 {day}일"
        backup_ws.cell(row=backup_row, column=1).font = Font(bold=True)
        backup_row += 1

        for r in rows:
            for c in range(1, 22):
                backup_ws.cell(row=backup_row, column=c).value = ws.cell(row=r, column=c).value
            backup_row += 1
        backup_row += 1

    if not messagebox.askyesno("확인", "백업 완료. 시트 데이터 삭제 및 코일교체 삽입을 진행할까요?"):
        wb.save(path)
        os.startfile(path)
        return

    DELETE_COL_LETTERS = ([chr(c) for c in range(ord('A'), ord('K') + 1)] + ['Q', 'R', 'T', 'U'])
    DELETE_COLS = [column_index_from_string(c) for c in DELETE_COL_LETTERS]

    for d in day_sheets:
        ws = wb[str(d)]
        for r in range(9, ws.max_row + 1):
            for c in DELETE_COLS:
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell, MergedCell): cell.value = None

    target_ws = wb["1"]
    found = False
    for r in range(backup_ws.max_row, 0, -1):
        if "코일교체" in str(backup_ws[f"B{r}"].value or ""):
            start, end = max(1, r - 1), min(backup_ws.max_row, r + 1)
            insert_row = 9
            for rr in range(start, end + 1):
                for c in range(1, 9):
                    target_ws.cell(row=insert_row, column=c).value = backup_ws.cell(row=rr, column=c).value
                insert_row += 1
            found = True
            break
    
    wb.save(path)
    messagebox.showinfo("완료", "월말 작업 완료.")
    os.startfile(path)

# ================= 메인 UI 클래스 =================
class ExcelCopyGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Automation v1.2")
        self.root.geometry("520x600")
        self.root.configure(bg="#f5f5f5")
        self.root.resizable(False, False)
        
        self.main_font = ("맑은 고딕", 11)
        self.title_font = ("맑은 고딕", 16, "bold")
        self.is_target_visible = False # 대상찾기 보임 상태 변수

        self.setup_style()
        self.create_widgets()

    def setup_style(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TLabel", background="#f5f5f5", font=self.main_font)
        style.configure("TButton", font=self.main_font, padding=5)
        style.configure("Header.TLabel", font=self.title_font, foreground="#333333")

    def create_widgets(self):
        # 상단 타이틀 & 백업 버튼
        top_frame = tk.Frame(self.root, bg="#f5f5f5")
        top_frame.pack(fill="x", padx=20, pady=15)
        ttk.Label(top_frame, text="엑셀 복사 자동화", style="Header.TLabel").pack(side="left")
        ttk.Button(top_frame, text="⚙ 백업", width=8, command=self.open_options).pack(side="right")

        # 메인 박스
        content = tk.Frame(self.root, bg="white", bd=1, relief="solid")
        content.pack(padx=20, pady=0, fill="both", expand=True)

        # 1. 원본 파일 선택
        self.add_section_title(content, "원본 파일 경로")
        f1_frame = ttk.Frame(content)
        f1_frame.pack(fill="x", padx=20, pady=5)
        self.file1 = ttk.Entry(f1_frame)
        self.file1.pack(side="left", fill="x", expand=True, ipady=4)
        ttk.Button(f1_frame, text="찾기", width=6, command=self.select_file1).pack(side="left", padx=(5, 0))

        # 2. 대상 파일 선택 (Toggle 기능)
        self.toggle_btn = tk.Label(content, text="▶ 대상 파일 수동 설정 (클릭)", bg="white", fg="#4a90e2", 
                                   font=("맑은 고딕", 9, "underline"), cursor="hand2")
        self.toggle_btn.pack(anchor="w", padx=20, pady=(10, 0))
        self.toggle_btn.bind("<Button-1>", lambda e: self.toggle_target_frame())

        self.target_frame = tk.Frame(content, bg="white") # 초기에는 숨김
        self.file2 = ttk.Entry(self.target_frame)
        self.file2.pack(side="left", fill="x", expand=True, ipady=4, padx=(20, 0))
        ttk.Button(self.target_frame, text="대상찾기", width=8, command=self.select_file2).pack(side="left", padx=(5, 20))

        # 3. 행 & 날짜 설정
        mid_frame = tk.Frame(content, bg="white")
        mid_frame.pack(fill="x", padx=20, pady=25)

        row_inner = tk.Frame(mid_frame, bg="white")
        row_inner.pack(side="left", expand=True)
        ttk.Label(row_inner, text="시작 행 위치").pack()
        self.start_row_entry = ttk.Entry(row_inner, width=10, justify="center")
        self.start_row_entry.insert(0, "9")
        self.start_row_entry.pack(pady=5)

        date_inner = tk.Frame(mid_frame, bg="white")
        date_inner.pack(side="left", expand=True)
        ttk.Label(date_inner, text="복사할 날짜(일)").pack()
        self.sheet = ttk.Combobox(date_inner, values=[str(i) for i in range(1, 32)], width=8, justify="center")
        self.sheet.set(str(datetime.today().day))
        self.sheet.pack(pady=5)

        # 4. 실행 버튼
        self.run_btn = tk.Button(self.root, text="복사 실행", bg="#4a90e2", fg="white",
                                 font=("맑은 고딕", 13, "bold"), relief="flat", cursor="hand2",
                                 command=self.run, pady=15)
        self.run_btn.pack(fill="x", padx=20, pady=25)
        self.run_btn.bind("<Enter>", lambda e: self.run_btn.configure(bg="#357abd"))
        self.run_btn.bind("<Leave>", lambda e: self.run_btn.configure(bg="#4a90e2"))

    def add_section_title(self, parent, text):
        lbl = ttk.Label(parent, text=text, font=("맑은 고딕", 9, "bold"), foreground="#666")
        lbl.pack(anchor="w", padx=20, pady=(15, 0))

    def toggle_target_frame(self):
        if self.is_target_visible:
            self.target_frame.pack_forget()
            self.toggle_btn.configure(text="▶ 대상 파일 수동 설정 (클릭)")
        else:
            self.target_frame.pack(fill="x", pady=5, after=self.toggle_btn)
            self.toggle_btn.configure(text="▼ 대상 파일 수동 설정 접기")
        self.is_target_visible = not self.is_target_visible

    def select_file1(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.file1.delete(0, tk.END); self.file1.insert(0, path)

    def select_file2(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.file2.delete(0, tk.END); self.file2.insert(0, path)

    def open_options(self):
        opt = tk.Toplevel(self.root)
        opt.title("백업 관리")
        opt.geometry("280x200")
        opt.grab_set()
        tk.Label(opt, text="월말 백업 실행", font=("맑은 고딕", 11, "bold")).pack(pady=15)
        v1, v6 = tk.BooleanVar(), tk.BooleanVar()
        tk.Checkbutton(opt, text="1호기 백업", variable=v1).pack()
        tk.Checkbutton(opt, text="6호기 백업", variable=v6).pack()
        ttk.Button(opt, text="실행", command=lambda: [run_backup("1호기") if v1.get() else None, 
                                                 run_backup("6호기") if v6.get() else None, opt.destroy()]).pack(pady=20)

    def detect_last_row(self, file_path):
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        panel_row = None
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=4).value
            if not val: continue
            text = str(val).replace(" ", "")
            if "H횡주관" in text: return r - 1, "H횡주관"
            if panel_row is None and "PANEL" in text.upper() and "TOTAL" in text.upper(): panel_row = r - 1
        return panel_row, "PANEL TOTAL" if panel_row else (None, None)

    def backup_target_file(self, target_path):
        backup_dir = r"D:\일지\백업"
        os.makedirs(backup_dir, exist_ok=True)
        base = os.path.basename(target_path)
        name, ext = os.path.splitext(base)
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        shutil.copy2(target_path, os.path.join(backup_dir, f"{name}_백업_{timestamp}{ext}"))
        backups = sorted([os.path.join(backup_dir, f) for f in os.listdir(backup_dir) if name in f], key=os.path.getmtime)
        while len(backups) > 5: os.remove(backups.pop(0))

    def run(self):
        if not self.file1.get():
            messagebox.showerror("오류", "원본 파일을 선택하세요")
            return

        # 대상 파일 결정 로직: 수동 입력값 있으면 사용, 없으면 자동 찾기
        target_path = self.file2.get()
        if not target_path:
            target_path = find_excel("1호기") # 바탕화면 자동찾기 우선
            if not target_path: # 바탕화면에 없으면 원본파일 폴더에서 찾기
                base_dir = os.path.dirname(self.file1.get())
                for file in os.listdir(base_dir):
                    if file.endswith(".xlsx") and "1호기" in file:
                        target_path = os.path.join(base_dir, file)
                        break

        if not target_path or not os.path.exists(target_path):
            messagebox.showerror("오류", "대상 파일을 찾을 수 없습니다. 수동으로 지정해주세요.")
            return

        try:
            start_row = int(self.start_row_entry.get())
            end_row, 기준 = self.detect_last_row(self.file1.get())
            if not end_row:
                messagebox.showerror("오류", "마지막 행 기준을 찾지 못했습니다.")
                return

            self.backup_target_file(target_path)
            df = pd.read_excel(self.file1.get(), header=None)
            wb2 = load_workbook(target_path)
            ws2 = wb2[self.sheet.get()]

            rows = end_row - 6
            for i in range(rows):
                ws2[f"J{start_row + i}"].value = df.iloc[6 + i, 3]
                ws2[f"I{start_row + i}"].value = df.iloc[6 + i, 4]
                ws2[f"U{start_row + i}"].value = df.iloc[6 + i, 9]

            wb2.save(target_path)
            messagebox.showinfo("완료", f"복사 완료! (기준: {기준})")
            os.startfile(target_path)
        except Exception as e:
            messagebox.showerror("오류", f"작업 중 에러: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    ExcelCopyGUI(root)
    root.mainloop()