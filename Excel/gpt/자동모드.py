import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string
import pandas as pd
import os
import shutil
import tkinter.font as tkFont
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")


# ==================================================
# 공통 함수
# ==================================================
def find_excel(machine_name):
    for f in os.listdir(DESKTOP):
        if f.endswith(".xlsx") and machine_name in f:
            return os.path.join(DESKTOP, f)
    return None


def prev_month():
    now = datetime.now()
    y, m = now.year, now.month - 0  # 현재달은 -0, 이전달은 -1,다음달은 +1
    if m == 0:
        y -= 1
        m = 12
    return y, m


# ==================================================
# 월말 백업 실행
# ==================================================
def run_backup(machine_name):
    path = find_excel(machine_name)
    if not path:
        messagebox.showerror("에러", f"{machine_name} 엑셀을 찾지 못했습니다.")
        return

    wb = load_workbook(path)

    # 날짜 시트
    day_sheets = [int(s) for s in wb.sheetnames if s.isdigit()]
    if not day_sheets:
        messagebox.showerror("에러", "날짜 시트가 없습니다.")
        return

    year, month = prev_month()
    backup_name = f"백업_{year}_{month:02d}"

    if backup_name in wb.sheetnames:
        messagebox.showwarning("중복", f"{backup_name} 시트가 이미 존재합니다.")
        return

    # ===== 백업 시트 생성 =====
    backup_ws = wb.create_sheet(backup_name)
    wb._sheets.remove(backup_ws)
    wb._sheets.append(backup_ws)

    backup_ws["A1"] = f"{year}년 {month}월 백업"
    backup_ws["A1"].font = Font(bold=True)

    backup_row = 3

    # ===== 날짜별 백업 =====
    for day in sorted(day_sheets):
        ws = wb[str(day)]
        rows = []

        for r in range(9, ws.max_row + 1):
            if ws[f"A{r}"].value in ("정미시간", "준비시간"):
                rows.append(r)

        if not rows:
            continue

        backup_ws.merge_cells(
            start_row=backup_row, start_column=1,
            end_row=backup_row, end_column=21
        )
        backup_ws.cell(row=backup_row, column=1).value = f"{month}월 {day}일"
        backup_ws.cell(row=backup_row, column=1).font = Font(bold=True)
        backup_row += 1

        for r in rows:
            for c in range(1, 22):
                backup_ws.cell(row=backup_row, column=c).value = ws.cell(row=r, column=c).value
            backup_row += 1

        backup_row += 1

        #print("백업 중:", month, "월", day, "일")  백업 테스트용입니다..

    # ===== 삭제 전 확인 =====
    if not messagebox.askyesno(
        "확인",
        "백업이 완료되었습니다.\n\n"
        "이제 날짜 시트(1~31)의 데이터를 삭제하고\n"
        "코일교체 내용을 1번 시트에 삽입할까요?"
    ):
        wb.save(path)
        messagebox.showinfo("완료", "백업만 완료되었습니다.\n엑셀을 엽니다.")
        os.startfile(path)
        return

    # ===== 날짜 시트 데이터 삭제 (선별 컬럼) =====
    DELETE_COL_LETTERS = (
        [chr(c) for c in range(ord('A'), ord('K') + 1)] +  # A~K
        ['Q', 'R'] +
        ['T', 'U']
    )

    DELETE_COLS = [column_index_from_string(c) for c in DELETE_COL_LETTERS]

    for d in day_sheets:
        ws = wb[str(d)]
        last_row = ws.max_row

        for r in range(9, last_row + 1):
            for c in DELETE_COLS:
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell, MergedCell):
                    cell.value = None

    # ===== 코일교체 → 1번 시트 =====
    backup_ws = wb[backup_name]
    target_ws = wb["1"]

    found = False

    for r in range(backup_ws.max_row, 0, -1):  # 아래 → 위
        cell_value = backup_ws[f"B{r}"].value

        if isinstance(cell_value, str) and "코일교체" in cell_value:
            start = max(1, r - 1)
            end = min(backup_ws.max_row, r + 1)

            insert_row = 9  # 1번 시트 A9부터

            for rr in range(start, end + 1):
                for c in range(1, 9):  # A~H
                    target_ws.cell(
                        row=insert_row,
                        column=c
                    ).value = backup_ws.cell(row=rr, column=c).value
                insert_row += 1

            found = True
            break   # ← 여기서 모든 작업 끝

    # 안전용 (이론상 안 뜸)
    if not found:
        messagebox.showwarning(
        "알림",
        "코일교체를 찾지 못했습니다.\n(데이터 구조를 확인해주세요)"
        )


    # ===== 전체 완료 =====
    wb.save(path)
    if messagebox.showinfo(
        "완료",
        f"{machine_name} 월말 작업이 모두 완료되었습니다.\n\n"
        "확인을 누르면 엑셀 파일을 엽니다."
    ):
        os.startfile(path)



# ================= 메인 UI 클래스 =================
class ExcelCopyGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Automation v1.2")
        self.root.geometry("500x520")
        self.root.configure(bg="#f5f5f5")
        
        # 폰트 설정
        self.main_font = ("맑은 고딕", 11)
        self.title_font = ("맑은 고딕", 16, "bold")
        
        self.setup_style()
        self.create_widgets()

    def setup_style(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TLabel", background="#f5f5f5", font=self.main_font)
        style.configure("TButton", font=self.main_font, padding=5)
        style.configure("Header.TLabel", font=self.title_font, foreground="#333333")

    def create_widgets(self):
        # 상단 타이틀 & 옵션 버튼
        top_frame = tk.Frame(self.root, bg="#f5f5f5")
        top_frame.pack(fill="x", padx=20, pady=10)
        
        ttk.Label(top_frame, text="엑셀 복사 자동화", style="Header.TLabel").pack(side="left")
        ttk.Button(top_frame, text="⚙ 백업", width=8, command=self.open_options).pack(side="right")

        # 메인 컨텐츠 영역
        content = tk.Frame(self.root, bg="white", bd=1, relief="solid")
        content.pack(padx=20, pady=10, fill="both", expand=True)

        # 1. 파일 선택
        self.add_section_title(content, "원본 파일 경로")
        f_frame = ttk.Frame(content, style="Card.TFrame")
        f_frame.pack(fill="x", padx=20, pady=5)
        
        self.file1 = ttk.Entry(f_frame)
        self.file1.pack(side="left", fill="x", expand=True, ipady=4)
        ttk.Button(f_frame, text="찾기", width=6, command=self.select_file1).pack(side="left", padx=(5, 0))

        # 2. 시작 행 & 날짜 (병렬 배치)
        mid_frame = tk.Frame(content, bg="white")
        mid_frame.pack(fill="x", padx=20, pady=20)

        # 시작 행
        row_inner = tk.Frame(mid_frame, bg="white")
        row_inner.pack(side="left", expand=True)
        ttk.Label(row_inner, text="시작 행 위치").pack()
        self.start_row_entry = ttk.Entry(row_inner, width=10, justify="center")
        self.start_row_entry.insert(0, "9")
        self.start_row_entry.pack(pady=5)

        # 날짜 선택
        date_inner = tk.Frame(mid_frame, bg="white")
        date_inner.pack(side="left", expand=True)
        ttk.Label(date_inner, text="복사할 날짜(일)").pack()
        self.sheet = ttk.Combobox(date_inner, values=[str(i) for i in range(1, 32)], width=8, justify="center")
        self.sheet.set(str(datetime.today().day))
        self.sheet.pack(pady=5)

        # 3. 실행 버튼
        self.run_btn = tk.Button(
            self.root, text="복사 작업 시작", bg="#4a90e2", fg="white",
            font=("맑은 고딕", 13, "bold"), relief="flat", cursor="hand2",
            command=self.run, pady=12
        )
        self.run_btn.pack(fill="x", padx=20, pady=25)
        
        # 버튼 호버 효과
        self.run_btn.bind("<Enter>", lambda e: self.run_btn.configure(bg="#357abd"))
        self.run_btn.bind("<Leave>", lambda e: self.run_btn.configure(bg="#4a90e2"))

    def add_section_title(self, parent, text):
        lbl = ttk.Label(parent, text=text, font=("맑은 고딕", 10, "bold"), foreground="#666")
        lbl.pack(anchor="w", padx=20, pady=(20, 0))

    def open_options(self):
        opt = tk.Toplevel(self.root)
        opt.title("월말 백업 설정")
        opt.geometry("280x220")
        opt.resizable(False, False)
        opt.grab_set() # 모달 창 설정

        tk.Label(opt, text="백업 대상 선택", font=("맑은 고딕", 11, "bold")).pack(pady=15)

        v1, v6 = tk.BooleanVar(), tk.BooleanVar()
        tk.Checkbutton(opt, text="1호기 월말 백업", variable=v1, font=self.main_font).pack(pady=5)
        tk.Checkbutton(opt, text="6호기 월말 백업", variable=v6, font=self.main_font).pack(pady=5)

        def run_selected():
            if not (v1.get() or v6.get()):
                return messagebox.showwarning("알림", "항목을 선택해주세요.")
            if v1.get(): run_backup("1호기")
            if v6.get(): run_backup("6호기")
            opt.destroy()

        ttk.Button(opt, text="백업 실행", command=run_selected).pack(pady=20)

    def select_file1(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.file1.delete(0, tk.END)
            self.file1.insert(0, path)
    
    # --------------------------------------------------
    def find_target_file(self, base_dir):
        for file in os.listdir(base_dir):
            if file.endswith(".xlsx") and "1호기" in file:
                return os.path.join(base_dir, file)
        return None

    # --------백업 폴더 생성------------------------------
    def backup_target_file(self, target_path):
        backup_dir = r"D:\일지\백업"
        os.makedirs(backup_dir, exist_ok=True)

        base = os.path.basename(target_path)
        name, ext = os.path.splitext(base)

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        backup_name = f"{name}_백업_{timestamp}{ext}"
        backup_path = os.path.join(backup_dir, backup_name)

        shutil.copy2(target_path, backup_path)

        backups = [
            os.path.join(backup_dir, f)
            for f in os.listdir(backup_dir)
            if f.endswith(".xlsx") and name in f
        ]
        backups.sort(key=os.path.getmtime)

        while len(backups) > 5:
            os.remove(backups.pop(0))

    # ------------자동 컷트 기준점, H횡준관 기준 & PANEL TOTAL 기준---------------
    def detect_last_row(self, file_path):
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        panel_row = None

        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=4).value
            if not val:
                continue

            text = str(val).replace(" ", "")

            if "H횡주관" in text:
                return r - 1, "H횡주관 기준"

            if panel_row is None and "PANEL" in text.upper() and "TOTAL" in text.upper():
                panel_row = r - 1

        if panel_row:
            return panel_row, "PANEL TOTAL 기준"

        return None, None

    # ---------------원본 파일 열수 없으면 새로운 Excel 창을 만듬니다--------------------
    def run(self):
        if not self.file1.get():
            messagebox.showerror("오류", "원본 파일을 선택하세요")
            return

        try:
            start_row = int(self.start_row_entry.get())
            if start_row < 1:
                raise ValueError
        except ValueError:
            messagebox.showerror("오류", "시작 행은 1 이상의 숫자여야 합니다")
            return

        base_dir = os.path.dirname(self.file1.get())

        # ===== 원본 읽기 =====
        try:
            df = pd.read_excel(self.file1.get(), header=None)
        except Exception:
            messagebox.showwarning(
                "원본 오류",
                "원본 Excel을 읽을 수 없습니다.\nExcel창을 열어드릴게요! 수동으로 복사하세요."
            )
            os.system("start excel")
            return

        # ===== 마지막 행 =====
        end_row, 기준 = self.detect_last_row(self.file1.get())
        if end_row is None:
            messagebox.showerror("오류", "마지막 행 기준을 찾지 못했습니다")
            return

        messagebox.showinfo("자동 판별", f"마지막 행은 [{기준}] 으로 결정되었습니다")

        # ===== 대상 파일 =====
        target_path = self.find_target_file(base_dir)
        if not target_path:
            messagebox.showerror(
                "대상 파일 없음",
                "바탕화면에서 '1호기'가 포함된 엑셀 파일을 찾지 못했습니다"
            )
            return

        # ===== 백업 =====
        self.backup_target_file(target_path)

        wb2 = load_workbook(target_path)
        ws2 = wb2[self.sheet.get()]

        start_df_row = 6
        rows = end_row - 6

        for i in range(rows):
            ws2[f"J{start_row + i}"].value = df.iloc[start_df_row + i, 3]
            ws2[f"I{start_row + i}"].value = df.iloc[start_df_row + i, 4]
            ws2[f"U{start_row + i}"].value = df.iloc[start_df_row + i, 9]

        try:
            wb2.save(target_path)
        except PermissionError:
            messagebox.showerror("오류", "대상 엑셀 파일이 열려 있습니다")
            return

        messagebox.showinfo(
            "완료",
            f"복사 완료\n\n대상 파일:\n{os.path.basename(target_path)}\n시작 행: {start_row}"
        )
        os.startfile(target_path)
        if not self.file1.get():
            return messagebox.showerror("오류", "원본 파일을 선택해주세요.")
        



# ================= 실행 =================
if __name__ == "__main__":
    root = tk.Tk()
    ExcelCopyGUI(root)
    root.mainloop()
