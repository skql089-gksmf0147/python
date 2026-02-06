import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font
from tkinter import Tk, Button, messagebox
from openpyxl.utils import column_index_from_string
from openpyxl.cell.cell import MergedCell

DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")


# ==================================================
# 공통 함수
# ==================================================
def find_excel(machine_name):
    for f in os.listdir(DESKTOP):
        if f.endswith(".xlsx") and machine_name in f:
            return os.path.join(DESKTOP, f)
    return None


def prev_month():
    now = datetime.now()
    y, m = now.year, now.month - 1
    if m == 0:
        y -= 1
        m = 12
    return y, m


# ==================================================
# 월말 백업 실행
# ==================================================
def run_backup(machine_name):
    path = find_excel(machine_name)
    if not path:
        messagebox.showerror("에러", f"{machine_name} 엑셀을 찾지 못했습니다.")
        return

    wb = load_workbook(path)

    # 날짜 시트
    day_sheets = [int(s) for s in wb.sheetnames if s.isdigit()]
    if not day_sheets:
        messagebox.showerror("에러", "날짜 시트가 없습니다.")
        return

    year, month = prev_month()
    backup_name = f"백업_{year}_{month:02d}"

    if backup_name in wb.sheetnames:
        messagebox.showwarning("중복", f"{backup_name} 시트가 이미 존재합니다.")
        return

    # ===== 백업 시트 생성 =====
    backup_ws = wb.create_sheet(backup_name)
    wb._sheets.remove(backup_ws)
    wb._sheets.append(backup_ws)

    backup_ws["A1"] = f"{year}년 {month}월 백업"
    backup_ws["A1"].font = Font(bold=True)

    backup_row = 3

    # ===== 날짜별 백업 =====
    for day in sorted(day_sheets):
        ws = wb[str(day)]
        rows = []

        for r in range(9, ws.max_row + 1):
            if ws[f"A{r}"].value in ("정미시간", "준비시간"):
                rows.append(r)

        if not rows:
            continue

        backup_ws.merge_cells(
            start_row=backup_row, start_column=1,
            end_row=backup_row, end_column=21
        )
        backup_ws.cell(row=backup_row, column=1).value = f"{month}월 {day}일"
        backup_ws.cell(row=backup_row, column=1).font = Font(bold=True)
        backup_row += 1

        for r in rows:
            for c in range(1, 22):
                backup_ws.cell(row=backup_row, column=c).value = ws.cell(row=r, column=c).value
            backup_row += 1

        backup_row += 1

        #print("백업 중:", month, "월", day, "일")

    # ===== 삭제 전 확인 =====
    if not messagebox.askyesno(
        "확인",
        "백업이 완료되었습니다.\n\n"
        "이제 날짜 시트(1~31)의 데이터를 삭제하고\n"
        "코일교체 내용을 1번 시트에 삽입할까요?"
    ):
        wb.save(path)
        messagebox.showinfo("완료", "백업만 완료되었습니다.\n엑셀을 엽니다.")
        os.startfile(path)
        return

    # ===== 날짜 시트 데이터 삭제 (선별 컬럼) =====
    DELETE_COL_LETTERS = (
        [chr(c) for c in range(ord('A'), ord('K') + 1)] +  # A~K
        ['Q', 'R'] +
        ['T', 'U']
    )

    DELETE_COLS = [column_index_from_string(c) for c in DELETE_COL_LETTERS]

    for d in day_sheets:
        ws = wb[str(d)]
        last_row = ws.max_row

        for r in range(9, last_row + 1):
            for c in DELETE_COLS:
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell, MergedCell):
                    cell.value = None

    # ===== 코일교체 → 1번 시트 =====
    backup_ws = wb[backup_name]
    target_ws = wb["1"]

    found = False

    for r in range(backup_ws.max_row, 0, -1):  # 아래 → 위
        cell_value = backup_ws[f"B{r}"].value

        if isinstance(cell_value, str) and "코일교체" in cell_value:
            start = max(1, r - 1)
            end = min(backup_ws.max_row, r + 1)

            insert_row = 9  # 1번 시트 A9부터

            for rr in range(start, end + 1):
                for c in range(1, 9):  # A~H
                    target_ws.cell(
                        row=insert_row,
                        column=c
                    ).value = backup_ws.cell(row=rr, column=c).value
                insert_row += 1

            found = True
            break   # ← 여기서 모든 작업 끝

    # 안전용 (이론상 안 뜸)
    if not found:
        messagebox.showwarning(
        "알림",
        "코일교체를 찾지 못했습니다.\n(데이터 구조를 확인해주세요)"
        )


    # ===== 전체 완료 =====
    wb.save(path)
    if messagebox.showinfo(
        "완료",
        f"{machine_name} 월말 작업이 모두 완료되었습니다.\n\n"
        "확인을 누르면 엑셀 파일을 엽니다."
    ):
        os.startfile(path)


# ==================================================
# GUI
# ==================================================
root = Tk()
root.title("월말 일지 백업")
root.geometry("420x260")

Button(
    root,
    text="1호기 월말 백업 실행",
    font=("맑은 고딕", 16),
    bg="#1E90FF",
    fg="white",
    command=lambda: run_backup("1호기")
).pack(pady=25)

Button(
    root,
    text="6호기 월말 백업 실행",
    font=("맑은 고딕", 16),
    bg="#2E8B57",
    fg="white",
    command=lambda: run_backup("6호기")
).pack(pady=10)

root.mainloop()
